import openpyxl as xl
import sys
import os


def word_finder(search_string, ws):
    result = []
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            try:
                v = ws.cell(i, j).value.lower()
                if v.find(search_string.lower()) != -1:
                    result.append(ws.cell(i, j))
            except:
                pass
    return result


def finder(search_string_input):
    filename = "list.xlsx"
    result = ""
    if filename.find("xlsx") != -1 and isinstance(search_string_input, str):

        try:
            wb = xl.load_workbook(filename)
            search_str = []
            not_found = True
            counter = 0

            for sheet in wb.worksheets:
                search_str = word_finder(search_string_input, sheet)
                for unit in search_str:
                    if hasattr(unit, 'value'):
                        result += f"{unit.value} - {sheet.title}<br>"
                        not_found = False
                        counter += 1
                    else:
                        pass
            if not_found:
                print(f"'{search_string_input}' wasn't found in file {filename}.")
            # else:
            #     result += f"Total {counter}."

            print(result)

        except:
            print("File not found")
    else:
        print("Invalid filename, try again.")


if __name__ == "__main__":
    search = str(sys.argv[1])
    finder(search)

